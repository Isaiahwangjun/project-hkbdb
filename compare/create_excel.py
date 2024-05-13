## 用來產生計算價格的空白 excel

import openpyxl


def create(name, root_path):
    file_list = [
        "BasicInformation", "Education", "Work", "Publication", "Article",
        "RelativeEvent", "Honor", "RelatedOrganizations", "Connections"
    ]

    # 創建一個新的 Excel 工作簿
    workbook = openpyxl.Workbook()

    # 添加 11 個表單
    for i in file_list:
        worksheet = workbook.create_sheet(title=i)

        # 添加欄位名稱
        worksheet['A1'] = 'key'
        worksheet['B1'] = 'GPT'
        worksheet['C1'] = 'HKBDB'
        worksheet['D1'] = 'score'

        # 一維表單不需要 'source'
        if i not in ["BasicInformation", "NameInformation", "Connections"]:
            worksheet['E1'] = 'source'

    # 移除預設的 Sheet
    workbook.remove(workbook['Sheet'])

    # 儲存 Excel 文件
    workbook.save(f'{root_path}/{name}/diff.xlsx')

    workbook = openpyxl.Workbook()

    # 添加 11 個表單
    for i in file_list:
        worksheet = workbook.create_sheet(title=i)

        # 添加欄位名稱
        worksheet['A1'] = 'key'
        worksheet['B1'] = 'value'

        if i not in ["BasicInformation", "NameInformation", "Connections"]:
            worksheet['C1'] = 'source'

    # 移除預設的 Sheet
    workbook.remove(workbook['Sheet'])

    # 儲存 Excel 文件
    workbook.save(f'{root_path}/{name}/GPThas.xlsx')
    workbook.save(f'{root_path}/{name}//HKBDBhas.xlsx')


def create_score(name, root_path):
    file_list = [
        "name", "BasicInformation", "Education", "Work", "Publication",
        "Article", "RelativeEvent", "Honor", "RelatedOrganizations",
        "Connections"
    ]

    # 創建一個新的 Excel 工作簿
    workbook = openpyxl.Workbook()

    accuracy_sheet = workbook.create_sheet(title='accuracy')
    onlyGPT_sheet = workbook.create_sheet(title='onlyGPT')
    onlyHKBDB_sheet = workbook.create_sheet(title='onlyHKBDB')

    for index, sheet_name in enumerate(file_list, start=1):
        accuracy_sheet.cell(row=1, column=index, value=sheet_name)
        onlyGPT_sheet.cell(row=1, column=index, value=sheet_name)
        onlyHKBDB_sheet.cell(row=1, column=index, value=sheet_name)

    # 移除預設的 Sheet
    workbook.remove(workbook['Sheet'])

    # 儲存 Excel 文件
    workbook.save(f'{root_path}/{name}/score.xlsx')
