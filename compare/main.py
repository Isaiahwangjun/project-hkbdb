import pandas as pd
from column_mapping import sheet_column_mapping
import oneDimension
import create_excel
import openpyxl
import sys
import os
from column_mapping import sheet_column_mapping


def fill0(name, root_path):
    # 將產生的準確率 excel 空白部分填 -1
    workbook = openpyxl.load_workbook(f'{root_path}/{name}/score.xlsx')

    # 遍历每个工作表
    for sheet_name in workbook.sheetnames:

        fillValue = 0
        if sheet_name == 'accuracy':
            fillValue = "-"

        # 获取当前工作表
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            # 检查每一行的每个单元格，如果为空则填充为0
            for cell in row:
                if cell.value is None:
                    cell.value = fillValue

    # 保存修改后的 Excel 文件
    workbook.save(f'{root_path}/{name}/score.xlsx')


def clear_excel(name, file, root_path):
    # delete empty sheets & not "keyword" data

    xls = pd.ExcelFile(f'{root_path}/{name}/{file}.xlsx')

    file_path = f'{root_path}/{name}/{file}-mod.xlsx'
    # if not os.path.exists(file_path):
    #     with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    #         pd.DataFrame().to_excel(writer, index=False, sheet_name='Sheet1')
    if os.path.exists(file_path):
        os.remove(file_path)
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, index=False, sheet_name='Sheet1')

    for sheet_name in xls.sheet_names:

        data = pd.read_excel(xls, sheet_name)
        if data.empty or len(data) == 0:
            continue

        specific_column = sheet_column_mapping.get(sheet_name)

        if isinstance(specific_column, list):
            if specific_column[0] is not None:
                if specific_column[0] in data.columns and len(data) > 0:
                    data.dropna(subset=[specific_column[0]], inplace=True)
                else:
                    data = pd.DataFrame()

        else:
            if specific_column is not None:
                if specific_column in data.columns and len(data) > 0:
                    data.dropna(subset=[specific_column], inplace=True)
                else:
                    data = pd.DataFrame()

        with pd.ExcelWriter(f'{root_path}/{name}/{file}-mod.xlsx',
                            mode='a',
                            engine='openpyxl',
                            if_sheet_exists='replace') as writer:
            data.to_excel(writer, sheet_name=sheet_name, index=False)


def read_excel_files(name, sheets, root_path):
    data = None
    ans = None
    try:
        data = pd.read_excel(f'{root_path}/{name}/gpt-mod.xlsx',
                             sheet_name=sheets)
    except:
        pass
    try:
        ans = pd.read_excel(f'{root_path}/{name}/hkbdb-mod.xlsx',
                            sheet_name=sheets)
    except:
        pass
    return data, ans


def process_sheets(name, sheets, root_path):

    # 判断要執行一維還是二維分析
    if sheets in sheet_column_mapping.keys():
        data, ans = read_excel_files(name, sheets, root_path)
        # print(sheets)
        if data is None:
            if ans is not None:
                oneDimension.onlyhas_two(ans, sheets, name, 'HKBDBhas',
                                         root_path)
        elif ans is None:
            oneDimension.onlyhas_two(data, sheets, name, 'GPThas', root_path)
        else:
            oneDimension.diffWith2File_two(data, ans, sheets, name, root_path)
    else:
        # 如果不在 sheet_column_mapping 中，則進行一維分析
        data, ans = read_excel_files(name, sheets, root_path)
        if data is None:
            if ans is not None:
                oneDimension.onlyhas(ans, sheets, name, 'HKBDBhas', root_path)
        elif ans is None:
            oneDimension.onlyhas(data, sheets, name, 'GPThas', root_path)
        else:
            oneDimension.diffWith2File_one(data, ans, sheets, name, root_path)


def main(name, root_path):
    # all_sheets = [
    #     "BasicInformation", "NameInformation", "Education", "Work",
    #     "Publication", "Article", "RelativeEvent", "Honor",
    #     "RelatedOrganizations", "Pieces", "Connections"
    # ]

    all_sheets = [
        "BasicInformation", "Education", "Work", "Publication", "Article",
        "RelativeEvent", "Honor", "RelatedOrganizations", "Connections"
    ]

    for sheets in all_sheets:
        process_sheets(name, sheets, root_path)


if __name__ == '__main__':

    root_path = f'C:/Users/wang/Desktop/daoyi/HongKong/capture100'
    name = sys.argv[1]
    clear_excel(name, 'gpt', root_path)
    clear_excel(name, 'hkbdb', root_path)

    create_excel.create(name, root_path)
    create_excel.create_score(name, root_path)

    main(name, root_path)
    # fill0(name, root_path)
    print(f"{name} done")
