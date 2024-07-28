import os
import json


def ddd():
    # 指定要遍历的目录路径
    directory = r'C:\Users\wang\Desktop\daoyi\HongKong\capture100'

    # 遍历目录及其子目录
    for root, dirs, files in os.walk(directory):
        for file in files:
            # 如果文件是xlsx文件
            if file.endswith('.xlsx'):
                # 构建原始文件的完整路径
                old_file_path = os.path.join(root, file)

                # 构建新文件的路径，可以根据需要修改成固定的文件名
                new_file_path = os.path.join(root, 'hkbdb.xlsx')

                # 重命名文件
                os.rename(old_file_path, new_file_path)
                print(f'Renamed {old_file_path} to {new_file_path}')


def organize_add():  #組織sheet 添加 '組織名稱' key
    # 指定根目录路径
    root_directory = r'C:\Users\wang\Desktop\daoyi\HongKong\capture100'

    # 遍历根目录及其子目录
    for root, dirs, files in os.walk(root_directory):
        for directory in dirs:
            # 构建当前文件夹中的 or.json 文件路径
            file_path = os.path.join(root, directory, 'organize.json')

            # 检查文件是否存在
            if os.path.exists(file_path):
                # 读取 JSON 数据
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)

            # 检查并修改数据
            for item in data['Related Organizations(相關組織)']:
                org_name = item.get(
                    '創立組織')  # 获取 '創立組織' 键对应的值，如果键不存在，则返回默认值 None
                if org_name and not org_name.isspace():  # 检查值是否存在并且非空
                    item['組織名稱'] = org_name
                else:
                    org_name = item.get(
                        '參與組織', None)  # 获取 '參與組織' 键对应的值，如果键不存在，则返回默认值 None
                    org_name = org_name.strip()
                    if org_name and not org_name.isspace():  # 检查值是否存在并且非空
                        item['組織名稱'] = org_name

            # 写入修改后的数据
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)
                print(f'Modified {file_path}')


from openpyxl import load_workbook


def change_sheet_name():

    root_directory = r'C:\Users\wang\Desktop\daoyi\HongKong\capture100'

    # 遍历根目录及其子目录
    for root, dirs, files in os.walk(root_directory):
        for directory in dirs:
            print(directory)
            # 构建当前文件夹中的 or.json 文件路径
            file_path = os.path.join(root, directory, 'hkbdb.xlsx')

            # 加载 Excel 文件
            workbook = load_workbook(file_path)

            # 定义一个字典，键是原始工作表名称，值是对应的新工作表名称
            sheet_name_mapping = {
                'person': 'BasicInformation',
                'educationevent': 'Education',
                'employmentevent': 'Work',
                'organizationevent': 'RelatedOrganizations',
                'publication': 'Publication',
                'Article1': 'Article',
                'event': 'RelativeEvent',
                'relationevent': 'Connections',
                'awardevent': 'Honor',
                'otherwork': 'Pieces'

                # 在这里继续添加其他工作表的名称映射
            }

            # 遍历字典，为每个工作表设置新的名称
            for old_name, new_name in sheet_name_mapping.items():
                # 检查工作表是否存在
                if old_name in workbook.sheetnames:
                    # 获取工作表对象
                    worksheet = workbook[old_name]
                    # 设置新的工作表名称
                    worksheet.title = new_name

            save_file_path = os.path.join(root, directory, 'hkbdb.xlsx')

            # 保存修改后的 Excel 文件
            workbook.save(save_file_path)


# 在 all_result.json 的表單名稱是舊的，轉為新的
def change2newKeyname(name):
    key_mapping = {
        "基本資料": "BasicInformation",
        "人名資料": "NameInformation",
        "學歷": "Education",
        "工作": "Work",
        "Publication(出版著作)": "Publication",
        "Article(單篇文章)": "Article",
        "其他作品": "Pieces",
        "Related Organizations(相關組織)": "RelatedOrganizations",
        "Connections(人際關係)": "Connections",
        "Relative Event(相關事件)": "RelativeEvent",
        "獎項": "Honor"
    }

    with open(
            f'C:/Users/wang/Desktop/daoyi/HongKong/capture100/{name}/all_result.json',
            'r',
            encoding='utf-8') as file:
        data = json.load(file)

    def rename_keys(data):

        new_data = {}
        for key, value in data.items():
            new_key = key_mapping.get(key, key)  # 如果存在映射关系，则使用映射后的键名，否则保持原来的键名
            if isinstance(value, dict):  # 如果值是字典类型，则递归调用该函数
                new_data[new_key] = rename_keys(value, key_mapping)
            else:
                new_data[new_key] = value
        return new_data

    new_data = rename_keys(data)

    with open(
            f'C:/Users/wang/Desktop/daoyi/HongKong/capture100/{name}/all_result.json',
            'w',
            encoding='utf-8') as file:
        json.dump(new_data, file, ensure_ascii=False, indent=4)


def hkbdb_add_organ_rdfs(
        name):  # hkbdb 下載的 excel organization 裡沒有 "rdfs:label"，新增此欄位以便後續計算
    import pandas as pd

    file_path = f'C:/Users/wang/Desktop/daoyi/HongKong/capture100/{name}/hkbdb.xlsx'
    sheet_name = 'RelatedOrganizations'
    xls = pd.ExcelFile(file_path)
    df_dict = {}
    for sheet_name in xls.sheet_names:
        df_dict[sheet_name] = pd.read_excel(xls, sheet_name)

    # 处理 RelatedOrganizations 工作表
    if 'RelatedOrganizations' in df_dict:
        df = df_dict['RelatedOrganizations']
        for index, row in df.iterrows():
            # 检查 hasFounded 或 hasParticipant 列是否包含特定值
            if 'hasFounded' in df.columns and pd.notnull(row['hasFounded']):
                df.at[index, 'rdfs:label'] = row['hasFounded']
            elif 'hasParticipant' in df.columns and pd.notnull(
                    row['hasParticipant']):
                df.at[index, 'rdfs:label'] = row['hasParticipant']
        # 将修改后的数据写回 RelatedOrganizations 工作表
        with pd.ExcelWriter(file_path) as writer:
            df.to_excel(writer, sheet_name='RelatedOrganizations', index=False)

    # 将其他工作表的数据写回到 Excel 文件中
    with pd.ExcelWriter(file_path,
                        mode='a',
                        engine='openpyxl',
                        if_sheet_exists='replace') as writer:
        for sheet_name, df in df_dict.items():
            if sheet_name != 'RelatedOrganizations':
                df.to_excel(writer, sheet_name=sheet_name, index=False)


def score_summary():

    root_directory = r'C:\Users\wang\Desktop\daoyi\HongKong\capture100'

    all_wb = load_workbook('summary4.xlsx', read_only=False)
    # 获取名为'accuracy'的工作表

    accuracy_ws = all_wb['accuracy']
    for root, dirs, files in os.walk(root_directory):
        for directory in dirs:
            # 构建当前文件夹中的 or.json 文件路径
            file_path = os.path.join(root, directory, 'score.xlsx')
            score_wb = load_workbook(file_path)
            # 获取名为'accuracy'的工作表的第二行数据
            accuracy_data = score_wb['accuracy'].iter_rows(min_row=2,
                                                           max_row=2,
                                                           values_only=True)
            # 将数据写入到all.xlsx的accuracy工作表中
            for row in accuracy_data:
                accuracy_ws.append(row)

    onlyGPT_ws = all_wb['onlyGPT']
    for root, dirs, files in os.walk(root_directory):
        for directory in dirs:
            # 构建当前文件夹中的 or.json 文件路径
            file_path = os.path.join(root, directory, 'score.xlsx')
            score_wb = load_workbook(file_path)
            # 获取名为'accuracy'的工作表的第二行数据
            accuracy_data = score_wb['onlyGPT'].iter_rows(min_row=2,
                                                          max_row=2,
                                                          values_only=True)
            # 将数据写入到all.xlsx的accuracy工作表中
            for row in accuracy_data:
                onlyGPT_ws.append(row)

    onlyGPTfull_ws = all_wb['onlyGPTfull']
    for root, dirs, files in os.walk(root_directory):
        for directory in dirs:
            # 构建当前文件夹中的 or.json 文件路径
            file_path = os.path.join(root, directory, 'score.xlsx')
            score_wb = load_workbook(file_path)
            # 获取名为'accuracy'的工作表的第二行数据
            accuracy_data = score_wb['onlyGPT'].iter_rows(min_row=3,
                                                          max_row=3,
                                                          values_only=True)
            # 将数据写入到all.xlsx的accuracy工作表中
            for row in accuracy_data:
                onlyGPTfull_ws.append(row)

    onlyHKBDB_ws = all_wb['onlyHKBDB']
    for root, dirs, files in os.walk(root_directory):
        for directory in dirs:
            # 构建当前文件夹中的 or.json 文件路径
            file_path = os.path.join(root, directory, 'score.xlsx')
            score_wb = load_workbook(file_path)
            # 获取名为'accuracy'的工作表的第二行数据
            accuracy_data = score_wb['onlyHKBDB'].iter_rows(min_row=2,
                                                            max_row=2,
                                                            values_only=True)
            # 将数据写入到all.xlsx的accuracy工作表中
            for row in accuracy_data:
                onlyHKBDB_ws.append(row)

    onlyHKBDBfull_ws = all_wb['onlyHKBDBfull']
    for root, dirs, files in os.walk(root_directory):
        for directory in dirs:
            # 构建当前文件夹中的 or.json 文件路径
            file_path = os.path.join(root, directory, 'score.xlsx')
            score_wb = load_workbook(file_path)
            # 获取名为'accuracy'的工作表的第二行数据
            accuracy_data = score_wb['onlyHKBDB'].iter_rows(min_row=3,
                                                            max_row=3,
                                                            values_only=True)
            # 将数据写入到all.xlsx的accuracy工作表中
            for row in accuracy_data:
                onlyHKBDBfull_ws.append(row)

    all_wb.save('summary4.xlsx')
    all_wb.close()


# import sys

# change2newKeyname(sys.argv[1])
# # print(sys.argv[1])
score_summary()
