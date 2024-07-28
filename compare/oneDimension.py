import pandas as pd
from diff_match_patch import diff_match_patch
import importlib
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from column_mapping import sheet_column_mapping
from fuzzywuzzy import fuzz
import DFappend
from score_mapping import sheet_score_mapping


def onlyhas(data, sheet, name, dcox, root_path):
    df = pd.DataFrame(columns=['key', 'value'])
    score_cnt = 0

    for _, row in data.iterrows():
        package = dynamic_import(sheet)
        for column in data.columns:
            if column in package.match_conditions:
                x = row[column]  # 从 data 中获取列的值
                if pd.notnull(x):
                    df = DFappend.onlyGPTorHKBDBhasDF(df, column, x)
                    score_cnt += 1

    mergeExcel(name, sheet, dcox, df, root_path)
    if dcox == 'HKBDBhas':
        writeInScoreExcel('onlyHKBDB', name, sheet, score_cnt, root_path)
    else:
        writeInScoreExcel('onlyGPT', name, sheet, score_cnt, root_path)


def onlyhas_two(data, sheet, name, dcox, root_path):
    df = pd.DataFrame(columns=['key', 'value', 'source'])
    sheet_columns = sheet_column_mapping.get(sheet, None)
    score_cnt = 0

    for _, row in data.iterrows():
        package = dynamic_import(sheet)
        if isinstance(sheet_columns, list):
            value = row.get(sheet_columns[0])
        else:
            value = row.get(sheet_columns)
        if value is None:
            value = 'None'
        else:
            score_cnt += 1
        for column in data.columns:
            if column in package.match_conditions:
                x = row[column]  # 从 data 中获取列的值
                if pd.notnull(x):
                    df = DFappend.onlyGPTorHKBDBhasDF(df, column, x, value)
    mergeExcel(name, sheet, dcox, df, root_path)
    if dcox == 'HKBDBhas':
        writeInScoreExcel('onlyHKBDB',
                          name,
                          sheet,
                          score_cnt,
                          root_path,
                          full=1)
    else:
        writeInScoreExcel('onlyGPT', name, sheet, score_cnt, root_path, full=1)


def dynamic_import(package_name):
    try:
        module = importlib.import_module(f"match_conditions.{package_name}")
        return module
    except ImportError:
        print(f"Failed to import module '{package_name}'")
        return None


def generate_diff(old_text, new_text):
    dmp = diff_match_patch()
    diffs = dmp.diff_main(old_text, new_text)
    dmp.diff_cleanupSemantic(diffs)
    diff_html = dmp.diff_prettyHtml(diffs)
    return diff_html


def mergeExcel(name, sheet, dcox, df, root_path):
    workbook = openpyxl.load_workbook(f'{root_path}/{name}/{dcox}.xlsx')
    worksheet = workbook[sheet]
    for row in dataframe_to_rows(df, index=False, header=False):
        worksheet.append(row)
    worksheet.append([' ', ' '])
    workbook.save(f'{root_path}/{name}/{dcox}.xlsx')


def writeInScoreExcel(type, name, sheet, score, root_path, full=None):
    workbook = openpyxl.load_workbook(f'{root_path}/{name}/score.xlsx')
    worksheet = workbook[type]

    if full is not None:
        name_write = f'{name}-Full'
    else:
        name_write = name

    specific_column_index = None
    for i, cell in enumerate(worksheet[1]):  # 第一行是标题行
        if cell.value == sheet:
            specific_column_index = i + 1

    for row in worksheet.iter_rows(min_row=2, max_col=specific_column_index):
        if row[0].value == name_write:
            worksheet.cell(row=row[0].row,
                           column=specific_column_index,
                           value=score)
            workbook.save(f'{root_path}/{name}/score.xlsx')
            return

    # 如果没有找到相同名称的行，则创建新行并添加分数
    max_row = worksheet.max_row
    worksheet.cell(row=max_row + 1, column=1, value=name_write)
    worksheet.cell(row=max_row + 1, column=specific_column_index, value=score)

    # 保存 Excel 文件
    workbook.save(f'{root_path}/{name}/score.xlsx')


def diffWith2File_one(data, ans, sheet, name, root_path):

    match_total_cnt = 0  # 計算兩者都有的總欄位數
    match_gpt_score = 0  # 計算兩者的相似度 or 準確率
    GPTmore = 0
    GPTless = 0
    hkbdbCnt = len(ans)

    with open(f'{root_path}/{name}/diff.html', "a", encoding='utf-8') as f:
        f.write(f"<h3>{sheet}</h3>")

    with open(f'{root_path}/{name}/same.html', "a", encoding='utf-8') as f:
        f.write(f"<h3>{sheet}</h3>")

    diff_df = pd.DataFrame(columns=['key', 'GPT', 'HKBDB', 'score', 'source'])
    GPThas_df = pd.DataFrame(columns=['key', 'value', 'source'])
    HKBDBhas_df = pd.DataFrame(columns=['key', 'value', 'source'])

    for index, row in data.iterrows():
        y_row = ans.iloc[index]  # 使用 index 將 y_row 從 ans 中選取對應行

        package = dynamic_import(sheet)
        # 對於每個匹配條件，根據其鍵選取相應的匹配函數
        for condition_key, match_function in package.match_conditions.items():

            x = row.get(condition_key)  # 從 data 中取得 x 的值
            y = y_row.get(condition_key)  # 從 ans 中取得 y 的值

            if (condition_key == 'hasPlaceOfBirth'
                    or 'hasNativePlace') and y is not None and pd.notnull(y):
                y = str(y).split('_')[0]

            if (condition_key == 'hasBirthDate'
                    or 'hasDeathDate') and y is not None and pd.notnull(y):
                y = str(y).replace('-', '')

            # 處理三種情況：data 有而 ans 沒有、ans 有而 data 沒有、兩者都有
            if x is not None and y is not None:

                # 都有的話，計算準確率
                match_result = match_function(str(x), str(y))
                # print(f"{condition_key} 匹配結果:", match_result)

                # total accuracy
                match_total_cnt += 1
                match_gpt_score += match_result / 100

                # 如果有差異，將差異的部分呈現 html &　excel
                if match_result != 100:
                    diff_html = generate_diff(str(x), str(y))
                    diff_html_with_info = f"{condition_key}: {diff_html} ({match_result})<br>"
                    with open(f'{root_path}/{name}/diff.html',
                              "a",
                              encoding='utf-8') as f:
                        f.write(diff_html_with_info)
                        f.write('\n')

                    diff_df = DFappend.diffDF(diff_df, condition_key, x, y,
                                              match_result)
                else:  #如果相似度 = 100, 網頁呈現
                    html_info = f"{condition_key} <span style='background-color: rgba(255, 0, 0, 0.3);'>{x} </span> <span style='background-color: rgba(0, 255, 0, 0.3);'>{y}</span><br>"
                    with open(f'{root_path}/{name}/same.html',
                              "a",
                              encoding='utf-8') as f:
                        f.write(html_info)
                        f.write('\n')

            elif x is not None and y is None:
                if not pd.isna(x):
                    # print(f"{condition_key} GPT有， HKDBD沒有")
                    GPThas_df = DFappend.onlyGPTorHKBDBhasDF(
                        GPThas_df, condition_key, x)
                    GPTmore += 1

            elif x is None and y is not None:
                if not pd.isna(y):
                    # print(f"{condition_key} HKBDB有， GPT沒有")
                    HKBDBhas_df = DFappend.onlyGPTorHKBDBhasDF(
                        HKBDBhas_df, condition_key, y)
                    GPTless += 1
    if match_total_cnt != 0:
        writeInScoreExcel('accuracy', name, sheet,
                          (match_gpt_score / match_total_cnt), root_path)
    else:
        print("one")
    writeInScoreExcel('onlyGPT', name, sheet, (GPTmore / hkbdbCnt), root_path)
    writeInScoreExcel('onlyHKBDB', name, sheet, (GPTless / hkbdbCnt),
                      root_path)

    mergeExcel(name, sheet, 'diff', diff_df, root_path)
    mergeExcel(name, sheet, 'GPThas', GPThas_df, root_path)
    mergeExcel(name, sheet, 'HKBDBhas', HKBDBhas_df, root_path)

    # match_accuracy = match_gpt_score / match_total_cnt
    # print(match_accuracy)


def diffWith2File_two(data, ans, sheet, name, root_path):

    match_total_cnt = 0  # 計算兩者都有的總欄位數
    match_gpt_score = 0  # 計算兩者的相似度 or 準確率
    GPTmore = 0
    GPTless = 0
    GPTmoreFull = 0
    GPTlessFull = 0

    with open(f"{root_path}/{name}/diff.html", "a", encoding='utf-8') as f:
        f.write(f"<h3>{sheet}</h3>")

    with open(f"{root_path}/{name}/same.html", "a", encoding='utf-8') as f:
        f.write(f"<h3>{sheet}</h3>")

    diff_df = pd.DataFrame(columns=['key', 'GPT', 'HKBDB', 'score', 'source'])
    GPThas_df = pd.DataFrame(columns=['key', 'value', 'source'])
    HKBDBhas_df = pd.DataFrame(columns=['key', 'value', 'source'])

    sheet_columns = sheet_column_mapping.get(sheet, None)
    EachCellScore = float(sheet_score_mapping.get(sheet, None))

    unmatched_data = pd.DataFrame()  # 存储未匹配到的 data 行
    unmatched_ans = ans.copy()  # 創建 ans 的副本用於追蹤未匹配的行

    # 如果拿到的是 list，代表要比較多個關鍵欄位
    for _, row_data in data.iterrows():
        best_match_ratio = -1  # 最佳匹配相似度
        best_match_row_ans = None  # 最佳匹配的 ans 行

        # 遍历 ans 的每一行
        for _, row_ans in unmatched_ans.iterrows():

            if isinstance(sheet_columns, list):
                # 计算当前行与 data 行的相似度
                if sheet_columns[0] in row_data and sheet_columns[0] in row_ans:
                    if row_data[sheet_columns[0]] is not None and row_ans[
                            sheet_columns[0]] is not None:
                        similarity_ratio = fuzz.ratio(
                            str(row_data[sheet_columns[0]]),
                            str(row_ans[sheet_columns[0]]))

                if sheet_columns[1] in row_data and sheet_columns[1] in row_ans:
                    if row_data[sheet_columns[1]] is not None and row_ans[
                            sheet_columns[1]] is not None:
                        similarity_ratio += fuzz.ratio(
                            str(row_data[sheet_columns[1]]),
                            str(row_ans[sheet_columns[1]]))
                # 更新最佳匹配
                if similarity_ratio > 100:
                    if similarity_ratio > best_match_ratio:
                        best_match_ratio = similarity_ratio
                        best_match_row_ans = row_ans
                        print(f"similarity_ratio:{similarity_ratio}")
                if unmatched_ans is not None:
                    continue
            else:
                for _, row_ans in unmatched_ans.iterrows():
                    # 计算当前行与 data 行的相似度
                    if sheet_columns in row_data and sheet_columns in row_ans:
                        if row_data[sheet_columns] is not None and row_ans[
                                sheet_columns] is not None:
                            similarity_ratio = fuzz.ratio(
                                str(row_data[sheet_columns]),
                                str(row_ans[sheet_columns]))

                            # 更新最佳匹配
                            if similarity_ratio > 50:
                                if similarity_ratio > best_match_ratio:
                                    best_match_ratio = similarity_ratio
                                    best_match_row_ans = row_ans

        # 如果找到了最佳匹配行，從未匹配的 ans 列表中刪除該行
        if best_match_row_ans is not None:
            unmatched_ans = unmatched_ans.drop(best_match_row_ans.name)
        else:
            unmatched_data = unmatched_data._append(row_data)
            continue

        package = dynamic_import(sheet)
        # 比较数据并存储差异信息
        for condition_key, match_function in package.match_conditions.items():
            x = row_data.get(condition_key)
            y = best_match_row_ans.get(condition_key)

            if ('Date' in condition_key) and y is not None and pd.notnull(y):
                y = str(y).replace('-', '')

            # 處理三種情況：data 有而 ans 沒有、ans 有而 data 沒有、兩者都有
            if pd.notnull(x) and pd.notnull(y):
                # 都有的话，计算准确率
                match_result = match_function(str(x), str(y))
                # print(f"{condition_key} {x} {y}  匹配結果:", match_result)

                # total accuracy
                match_total_cnt += 1
                match_gpt_score += match_result / 100
                if sheet == 'RelativeEvent':
                    print(x, y, match_result)

                # 如果有差異，將差異的部分存入 diff_df
                if match_result != 100:
                    diff_html = generate_diff(str(x), str(y))
                    diff_html_with_info = f"{condition_key}: {diff_html} ({match_result})<br>"
                    with open(f"{root_path}/{name}/diff.html",
                              "a",
                              encoding='utf-8') as f:
                        f.write(diff_html_with_info)
                        f.write('\n')
                    if isinstance(sheet_columns, list):
                        value = best_match_row_ans.get(sheet_columns[0] +
                                                       sheet_columns[1])
                    else:
                        value = best_match_row_ans.get(sheet_columns)
                    diff_df = DFappend.diffDF(diff_df, condition_key, x, y,
                                              match_result, value)

                else:  #如果相似度 = 100, 網頁呈現
                    html_info = f"{condition_key} <span style='background-color: rgba(255, 0, 0, 0.3);'>{x} </span> <span style='background-color: rgba(0, 255, 0, 0.3);'>{y}</span><br>"
                    with open(f"{root_path}/{name}/same.html",
                              "a",
                              encoding='utf-8') as f:
                        f.write(html_info)
                        f.write('\n')

            #4/29 接著做 data 有而 ans 沒有、ans 有而 data 沒有
            elif pd.notnull(x):
                # GPTmore += EachCellScore
                GPTmore += 1
                if isinstance(sheet_columns, list):
                    value = best_match_row_ans.get(sheet_columns[0] +
                                                   sheet_columns[1])
                else:
                    value = best_match_row_ans.get(sheet_columns)
                if value is None:
                    value = 'None'
                GPThas_df = DFappend.onlyGPTorHKBDBhasDF(
                    GPThas_df, condition_key, x, value)

            elif pd.notnull(y):
                # GPTless += EachCellScore
                GPTless += 1
                if isinstance(sheet_columns, list):
                    value = best_match_row_ans.get(sheet_columns[0] +
                                                   sheet_columns[1])
                else:
                    value = best_match_row_ans.get(sheet_columns)
                if value is None:
                    value = 'None'
                HKBDBhas_df = DFappend.onlyGPTorHKBDBhasDF(
                    HKBDBhas_df, condition_key, y, value)

    for _, row_data in unmatched_data.iterrows():
        # GPTmore += 1
        GPTmoreFull += 1
        package = dynamic_import(sheet)
        for condition_key, match_function in package.match_conditions.items():
            x = row_data.get(condition_key)
            if pd.notnull(x):
                value = row_data.get(sheet_columns)
                if value is None:
                    value = 'None'
                GPThas_df = DFappend.onlyGPTorHKBDBhasDF(
                    GPThas_df, condition_key, x, value)

    for _, row_data in unmatched_ans.iterrows():
        # GPTless += 1
        GPTlessFull += 1
        package = dynamic_import(sheet)
        for condition_key, match_function in package.match_conditions.items():
            x = row_data.get(condition_key)
            if pd.notnull(x):
                value = row_data.get(sheet_columns)
                if value is None:
                    value = 'None'
                HKBDBhas_df = DFappend.onlyGPTorHKBDBhasDF(
                    HKBDBhas_df, condition_key, x, value)

    if match_total_cnt != 0:
        writeInScoreExcel('accuracy', name, sheet,
                          (match_gpt_score / match_total_cnt), root_path)
    else:
        print("two")
    # writeInScoreExcel('onlyGPT', name, sheet, (GPTmore / hkbdbCnt), root_path)
    # writeInScoreExcel('onlyHKBDB', name, sheet, (GPTless / hkbdbCnt),
    #                   root_path)

    writeInScoreExcel('onlyGPT', name, sheet, GPTmore, root_path)
    writeInScoreExcel('onlyHKBDB', name, sheet, GPTless, root_path)

    writeInScoreExcel('onlyGPT', name, sheet, GPTmoreFull, root_path, full=1)
    writeInScoreExcel('onlyHKBDB', name, sheet, GPTlessFull, root_path, full=1)

    # mergeExcel(name, sheet, 'diff', diff_df, root_path)
    # mergeExcel(name, sheet, 'GPThas', GPThas_df, root_path)
    # mergeExcel(name, sheet, 'HKBDBhas', HKBDBhas_df, root_path)
